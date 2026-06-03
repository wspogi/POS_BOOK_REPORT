import os
import calendar
import logging
from datetime import date, datetime, timedelta

import pandas as pd
import pyodbc
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


# ============================================================
# POS Book Report Automation
# ============================================================
#
# Purpose:
#   This script generates the monthly POS Book Report from
#   the POS SQL Server database.
#
# What it does:
#   1. Reads database credentials and report settings from .env.
#   2. Connects to the POS SQL Server database.
#   3. Extracts POS sales invoice / receipt data.
#   4. Parses customer details from HISTSUB.
#   5. Parses VAT breakdown from HISTSUB.
#   6. Maps Senior / PWD / Other discounts based on disccode.
#   7. Writes the result to the Excel template.
#   8. Saves the generated report to the output folder.
#
# Confirmed POS source tables:
#   HISTMAIN = receipt/header table
#   HISTSUB  = transaction details, customer details, VAT breakdown
#   MDISC    = discount master reference
#   TTRNJRNL = printed receipt text / journal, used only for checking
#
# Confirmed discount mapping from MDISC:
#   05 = SENIOR DISC
#   A1 = PWD DISCOUNT
#   06 = SPECIAL, mapped to OTHERS
#   01 = EDIS, mapped to OTHERS
#   07 = OPEN AMT, mapped to OTHERS
#
# Expected Excel template:
#   Row 1 to 7 = report title/details
#   Row 8      = column headers
#   Row 9+     = data rows
#
# ============================================================


# ============================================================
# Base paths
# ============================================================

# BASE_DIR means the folder where this main.py file is located.
# Example:
#   C:\Users\YourName\Desktop\POS_BOOK_REPORT
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ENV_PATH points to the .env file inside the same project folder.
ENV_PATH = os.path.join(BASE_DIR, ".env")

# Load .env values so Python can read them using os.getenv().
load_dotenv(ENV_PATH)


# ============================================================
# Report columns
# ============================================================
#
# These columns must match the Excel POS Book Report format.
# The order here is also the order used when writing data to Excel.

REPORT_COLUMNS = [
    "DATE",
    "OR NO",
    "ID NO",
    "NAME",
    "ADDRESS",
    "TIN NO",
    "TOTAL SALES",
    "VATABLE SALES",
    "VAT EXEMPT",
    "ZERO RATED",
    "VAT 12%",
    "SENIOR",
    "PWD",
    "OTHERS",
    "NET SALES",
]


# ============================================================
# Environment helpers
# ============================================================

def get_required_env(name: str) -> str:
    """
    Reads a required value from .env.

    Example:
        SQL_SERVER=192.168.254.109

    If the value is missing or blank, the script stops immediately.
    This helps us detect wrong or incomplete .env setup.
    """
    value = os.getenv(name)

    if value is None or value.strip() == "":
        raise ValueError(f"Missing required .env value: {name}")

    return value.strip()


def get_optional_env(name: str, default_value: str) -> str:
    """
    Reads an optional value from .env.

    If the value does not exist, it returns the default value.

    Example:
        OUTPUT_DIR=output

    If OUTPUT_DIR is missing, it will use:
        output
    """
    value = os.getenv(name)

    if value is None or value.strip() == "":
        return default_value

    return value.strip()


# ============================================================
# Logging setup
# ============================================================

def setup_logging(log_dir: str):
    """
    Creates logging setup for the automation.

    Logs will be saved inside the LOG_DIR folder.

    Example log file:
        logs/POS_BOOK_REPORT_20260603_143000.log

    This is useful kapag naka Task Scheduler na siya,
    kasi kahit hindi mo makita yung CMD window, may log file pa rin.
    """
    log_folder_path = os.path.join(BASE_DIR, log_dir)
    os.makedirs(log_folder_path, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(
        log_folder_path,
        f"POS_BOOK_REPORT_{timestamp}.log",
    )

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )

    logging.info("Logging started.")
    logging.info("Log file: %s", log_file)

    return log_file


# ============================================================
# SQL Server connection
# ============================================================

def get_connection():
    """
    Creates SQL Server connection using credentials from .env.

    Required .env values:
        SQL_SERVER
        SQL_DATABASE
        SQL_USERNAME
        SQL_PASSWORD

    Notes:
        TrustServerCertificate=yes is included to avoid certificate
        trust issues when connecting to local/network SQL Server.
    """
    server = get_required_env("SQL_SERVER")
    database = get_required_env("SQL_DATABASE")
    username = get_required_env("SQL_USERNAME")
    password = get_required_env("SQL_PASSWORD")

    conn_str = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={username};"
        f"PWD={password};"
        "TrustServerCertificate=yes;"
    )

    logging.info("Connecting to SQL Server...")
    logging.info("SQL_SERVER: %s", server)
    logging.info("SQL_DATABASE: %s", database)
    logging.info("SQL_USERNAME: %s", username)

    return pyodbc.connect(conn_str, timeout=30)


# ============================================================
# Date range logic
# ============================================================

def get_report_date_range():
    """
    Builds report date range based on DATE_MODE from .env.

    Supported DATE_MODE values:

    1. DATE_MODE=PREVIOUS_MONTH
       Automatically gets the previous calendar month.

       Example:
         If today is 2026-06-03,
         report period will be:
           2026-05-01 to 2026-05-31

       SQL date range will be:
           HM.trandate >= '2026-05-01'
           HM.trandate <  '2026-06-01'

    2. DATE_MODE=CUSTOM
       Uses DATE_FROM and DATE_TO from .env.

       Example .env:
           DATE_MODE=CUSTOM
           DATE_FROM=2026-04-01
           DATE_TO=2026-04-30

       Python converts DATE_TO to exclusive SQL date:
           SQL Date From = 2026-04-01
           SQL Date To   = 2026-05-01

    Why DATE_TO is exclusive:
        This avoids missing transactions with time values.
        Best practice for SQL date filtering is:
            trandate >= date_from
            trandate < next_day_after_date_to
    """
    date_mode = get_optional_env("DATE_MODE", "PREVIOUS_MONTH").upper()

    if date_mode == "PREVIOUS_MONTH":
        today = date.today()

        # First day of current month.
        # Example: 2026-06-01
        first_day_current_month = date(today.year, today.month, 1)

        # Last day of previous month.
        # Example: 2026-05-31
        last_day_previous_month = first_day_current_month - timedelta(days=1)

        # First day of previous month.
        # Example: 2026-05-01
        date_from = date(
            last_day_previous_month.year,
            last_day_previous_month.month,
            1,
        )

        # Exclusive end date for SQL.
        # Example: 2026-06-01
        date_to = first_day_current_month

        report_year = date_from.year
        report_month = date_from.month

        return date_from, date_to, report_year, report_month, date_mode

    if date_mode == "CUSTOM":
        raw_date_from = get_required_env("DATE_FROM")
        raw_date_to = get_required_env("DATE_TO")

        # DATE_FROM from .env.
        date_from = datetime.strptime(raw_date_from, "%Y-%m-%d").date()

        # DATE_TO from .env is treated as inclusive.
        date_to_inclusive = datetime.strptime(raw_date_to, "%Y-%m-%d").date()

        # Add 1 day so SQL can use < date_to.
        date_to = date_to_inclusive + timedelta(days=1)

        report_year = date_from.year
        report_month = date_from.month

        return date_from, date_to, report_year, report_month, date_mode

    raise ValueError("Invalid DATE_MODE. Use PREVIOUS_MONTH or CUSTOM.")


# ============================================================
# SQL query
# ============================================================

def get_pos_book_query() -> str:
    """
    Final POS Book Report SQL query.

    Important source logic:

    1. HISTMAIN
       Main receipt/header table.
       One row per OR/receipt.

       Used fields:
           receipt
           trandate
           amount
           disccode
           discamt
           trantype

    2. HISTSUB customer row
       Customer details are stored in HISTSUB where:
           type = 'M'
           code = 'CT'

       Sample raw value:
           þREYþþþþþTAGASþ0þþþþ0

       Parsed positions:
           RowNo 2 = Customer Name
           RowNo 7 = Address
           RowNo 8 = TIN

    3. HISTSUB VAT row
       VAT details are stored in HISTSUB where:
           type = 'U'
           code = '12'

       Sample raw value:
           12þ16082.14þ1929.86þ846.43þþ16928.57

       Parsed positions:
           RowNo 2 = VATable Sales
           RowNo 3 = VAT Amount
           RowNo 6 = Net Sales

    4. Discount mapping
       Based on HISTMAIN.disccode:
           05 = Senior
           A1 = PWD
           Others = all other non-blank discount codes
    """
    return """
WITH Base AS
(
    SELECT
        HM.transact,
        HM.receipt,
        HM.trandate,
        HM.amount,
        HM.disccode,
        HM.discamt,

        CAST(CT.reference AS NVARCHAR(MAX)) AS CustomerRaw,
        CAST(VAT.reference AS NVARCHAR(MAX)) AS VatRaw
    FROM dbo.HISTMAIN HM

    LEFT JOIN dbo.HISTSUB CT
        ON HM.transact = CT.transact
        AND CT.type = 'M'
        AND LTRIM(RTRIM(CT.code)) = 'CT'

    LEFT JOIN dbo.HISTSUB VAT
        ON HM.transact = VAT.transact
        AND VAT.type = 'U'
        AND LTRIM(RTRIM(VAT.code)) = '12'

    WHERE HM.trandate >= ?
      AND HM.trandate < ?
      AND HM.trantype = 'A'
),
Parsed AS
(
    SELECT
        B.*,

        CTParsed.CustomerName,
        CTParsed.CustomerAddress,
        CTParsed.CustomerTIN,

        VATParsed.VatableSales,
        VATParsed.VatAmount,
        VATParsed.NetSales

    FROM Base B

    OUTER APPLY
    (
        SELECT
            MAX(CASE WHEN X.RowNo = 2 THEN X.Value END) AS CustomerName,
            MAX(CASE WHEN X.RowNo = 7 THEN X.Value END) AS CustomerAddress,
            MAX(CASE WHEN X.RowNo = 8 THEN X.Value END) AS CustomerTIN
        FROM
        (
            SELECT
                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNo,
                LTRIM(RTRIM(T.C.value('.', 'NVARCHAR(MAX)'))) AS Value
            FROM
            (
                SELECT TRY_CAST(
                    '<x><i>' +
                    REPLACE(
                        REPLACE(
                            REPLACE(ISNULL(B.CustomerRaw, ''), '&', '&amp;'),
                            '<', '&lt;'
                        ),
                        N'þ',
                        '</i><i>'
                    )
                    + '</i></x>' AS XML
                ) AS XmlData
            ) D
            CROSS APPLY D.XmlData.nodes('/x/i') T(C)
        ) X
    ) CTParsed

    OUTER APPLY
    (
        SELECT
            TRY_CONVERT(MONEY, MAX(CASE WHEN X.RowNo = 2 THEN X.Value END)) AS VatableSales,
            TRY_CONVERT(MONEY, MAX(CASE WHEN X.RowNo = 3 THEN X.Value END)) AS VatAmount,
            TRY_CONVERT(MONEY, MAX(CASE WHEN X.RowNo = 6 THEN X.Value END)) AS NetSales
        FROM
        (
            SELECT
                ROW_NUMBER() OVER (ORDER BY (SELECT NULL)) AS RowNo,
                LTRIM(RTRIM(T.C.value('.', 'NVARCHAR(MAX)'))) AS Value
            FROM
            (
                SELECT TRY_CAST(
                    '<x><i>' +
                    REPLACE(
                        REPLACE(
                            REPLACE(ISNULL(B.VatRaw, ''), '&', '&amp;'),
                            '<', '&lt;'
                        ),
                        N'þ',
                        '</i><i>'
                    )
                    + '</i></x>' AS XML
                ) AS XmlData
            ) D
            CROSS APPLY D.XmlData.nodes('/x/i') T(C)
        ) X
    ) VATParsed
)
SELECT
    CAST(trandate AS DATE) AS [DATE],
    receipt AS [OR NO],

    '' AS [ID NO],

    ISNULL(CustomerName, '') AS [NAME],
    ISNULL(CustomerAddress, '') AS [ADDRESS],
    ISNULL(CustomerTIN, '') AS [TIN NO],

    amount AS [TOTAL SALES],

    ISNULL(VatableSales, ROUND(amount / 1.12, 2)) AS [VATABLE SALES],

    CAST(0.00 AS MONEY) AS [VAT EXEMPT],
    CAST(0.00 AS MONEY) AS [ZERO RATED],

    ISNULL(VatAmount, ROUND(amount - (amount / 1.12), 2)) AS [VAT 12%],

    CASE 
        WHEN LTRIM(RTRIM(disccode)) = '05' THEN discamt
        ELSE CAST(0.00 AS MONEY)
    END AS [SENIOR],

    CASE 
        WHEN LTRIM(RTRIM(disccode)) = 'A1' THEN discamt
        ELSE CAST(0.00 AS MONEY)
    END AS [PWD],

    CASE 
        WHEN LTRIM(RTRIM(disccode)) NOT IN ('', '05', 'A1') THEN discamt
        ELSE CAST(0.00 AS MONEY)
    END AS [OTHERS],

    ISNULL(NetSales, ISNULL(VatableSales, ROUND(amount / 1.12, 2))) AS [NET SALES]

FROM Parsed
ORDER BY trandate, receipt;
"""


# ============================================================
# Fetch data
# ============================================================

def fetch_report_data(conn, date_from: date, date_to: date) -> pd.DataFrame:
    """
    Runs the POS Book Report SQL query and returns the result
    as a pandas DataFrame.

    Parameters:
        conn      = active SQL Server connection
        date_from = SQL start date, inclusive
        date_to   = SQL end date, exclusive
    """
    query = get_pos_book_query()

    logging.info("Running POS Book Report SQL query...")
    logging.info("Query date_from: %s", date_from)
    logging.info("Query date_to exclusive: %s", date_to)

    df = pd.read_sql_query(
        query,
        conn,
        params=[date_from, date_to],
    )

    # Ensure the DataFrame follows the exact Excel column order.
    df = df[REPORT_COLUMNS]

    logging.info("Rows fetched: %s", len(df))

    return df


# ============================================================
# Excel helpers
# ============================================================

def clear_old_data(ws, start_row: int, end_col: int):
    """
    Clears old report rows from the Excel template.

    This keeps title/header rows intact.

    Example:
        start_row = 9
        end_col   = 15

    It clears:
        A9:O<last row>
    """
    max_row = ws.max_row

    for row in range(start_row, max_row + 1):
        for col in range(1, end_col + 1):
            ws.cell(row=row, column=col).value = None


def apply_basic_format(ws, start_row: int, end_row: int, end_col: int):
    """
    Applies simple formatting to generated rows.

    Formatting included:
        - Thin border
        - Vertical center alignment
        - Date format for DATE column
        - Number format for money columns
    """
    thin_side = Side(style="thin")

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    for row in range(start_row, end_row + 1):
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # DATE column = column A
    # Money columns = columns G to O
    money_columns = range(7, 16)

    for row in range(start_row, end_row + 1):
        ws.cell(row=row, column=1).number_format = "mm/dd/yyyy"

        for col in money_columns:
            ws.cell(row=row, column=col).number_format = "#,##0.00"


def write_to_excel_template(
    df: pd.DataFrame,
    template_file: str,
    output_file: str,
    branch: str,
    report_year: int,
    report_month: int,
):
    """
    Writes the DataFrame result to the Excel template.

    Parameters:
        df           = report data
        template_file = Excel template path
        output_file   = final Excel output path
        branch        = branch code/name from .env
        report_year   = report year
        report_month  = report month

    Template assumption:
        Row 1 to 7 = title/details
        Row 8      = headers
        Row 9+     = data rows
    """
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"Template file not found: {template_file}")

    logging.info("Opening Excel template: %s", template_file)

    wb = load_workbook(template_file)
    ws = wb.active

    start_row = 9
    end_col = len(REPORT_COLUMNS)

    # Build period label for Excel header.
    last_day = calendar.monthrange(report_year, report_month)[1]
    period_from = date(report_year, report_month, 1)
    period_to = date(report_year, report_month, last_day)

    # Update report header lines.
    # Adjust these cells if your Excel template uses different positions.
    ws["A3"] = "POS BOOK REPORT"
    ws["A4"] = (
        f"For the Period of "
        f"{period_from.strftime('%m/%d/%Y')} "
        f"to {period_to.strftime('%m/%d/%Y')}"
    )
    ws["A5"] = branch

    # Clear old generated rows from the template.
    clear_old_data(ws, start_row=start_row, end_col=end_col)

    # Write report rows to Excel.
    current_row = start_row

    for _, record in df.iterrows():
        for col_index, column_name in enumerate(REPORT_COLUMNS, start=1):
            value = record[column_name]

            # Convert DATE value to Excel-compatible datetime.
            if column_name == "DATE" and pd.notna(value):
                if isinstance(value, pd.Timestamp):
                    value = value.to_pydatetime()
                elif isinstance(value, date):
                    value = datetime(value.year, value.month, value.day)

            # Convert pandas NaN/NaT to blank cell.
            if pd.isna(value):
                value = None

            ws.cell(row=current_row, column=col_index).value = value

        current_row += 1

    # Apply formatting only if there are rows.
    if len(df) > 0:
        apply_basic_format(
            ws,
            start_row=start_row,
            end_row=current_row - 1,
            end_col=end_col,
        )

    # Make sure output folder exists.
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    logging.info("Saving Excel output: %s", output_file)
    wb.save(output_file)


# ============================================================
# Main process
# ============================================================

def main():
    """
    Main automation process.

    Required .env sample:

        SQL_SERVER=192.168.xxx.xxx
        SQL_DATABASE=VQPBOS
        SQL_USERNAME=your_username
        SQL_PASSWORD=your_password_here

        POS_BRANCH=CAM

        DATE_MODE=PREVIOUS_MONTH
        # DATE_FROM=2026-04-01
        # DATE_TO=2026-04-30

        TEMPLATE_FILE=templates/APR.xlsx

        OUTPUT_DIR=output
        LOG_DIR=logs
    """
    branch = get_required_env("POS_BRANCH")

    # New date handling:
    # This no longer needs REPORT_YEAR and REPORT_MONTH from .env.
    date_from, date_to, report_year, report_month, date_mode = get_report_date_range()

    template_file = get_required_env("TEMPLATE_FILE")
    output_dir = get_optional_env("OUTPUT_DIR", "output")
    log_dir = get_optional_env("LOG_DIR", "logs")

    # Start logging after LOG_DIR is known.
    setup_logging(log_dir)

    template_path = os.path.join(BASE_DIR, template_file)
    output_folder_path = os.path.join(BASE_DIR, output_dir)

    output_filename = f"POS_BOOK_REPORT_{branch}_{report_year}_{report_month:02d}.xlsx"
    output_file = os.path.join(output_folder_path, output_filename)

    logging.info("==========================================")
    logging.info("POS Book Report Automation")
    logging.info("==========================================")
    logging.info("Branch: %s", branch)
    logging.info("Date Mode: %s", date_mode)
    logging.info("Report Year: %s", report_year)
    logging.info("Report Month: %02d", report_month)
    logging.info("SQL Date From: %s", date_from)
    logging.info("SQL Date To Exclusive: %s", date_to)
    logging.info("Template: %s", template_path)
    logging.info("Output: %s", output_file)
    logging.info("==========================================")

    conn = None

    try:
        conn = get_connection()

        logging.info("Connected to SQL Server.")
        logging.info("Fetching POS Book Report data...")

        df = fetch_report_data(conn, date_from, date_to)

        logging.info("Writing Excel report...")

        write_to_excel_template(
            df=df,
            template_file=template_path,
            output_file=output_file,
            branch=branch,
            report_year=report_year,
            report_month=report_month,
        )

        logging.info("POS Book Report generation completed successfully.")
        logging.info("Generated file: %s", output_file)

    except Exception as e:
        logging.exception("Failed to generate POS Book Report.")
        raise

    finally:
        if conn is not None:
            conn.close()
            logging.info("SQL Server connection closed.")


if __name__ == "__main__":
    main()